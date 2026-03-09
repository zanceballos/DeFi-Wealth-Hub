from __future__ import annotations

from pathlib import Path


def parse_xlsx(file_path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for xlsx parsing") from exc

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []

    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            continue

        header = [str(item).strip() if item is not None else "" for item in values[0]]
        for record in values[1:]:
            payload: dict[str, str] = {}
            for index, header_name in enumerate(header):
                if not header_name:
                    continue
                value = record[index] if index < len(record) else None
                payload[header_name] = "" if value is None else str(value).strip()

            if any(value for value in payload.values()):
                rows.append(payload)

    workbook.close()
    return rows
